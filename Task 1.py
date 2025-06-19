from openpyxl import load_workbook

wb = load_workbook('sagatave_eksamenam.xlsx', data_only=True)

# Print available sheet names to confirm the correct one
print("Available sheets:", wb.sheetnames)

# Use the first sheet (or change index if needed)
ws = wb[wb.sheetnames[0]]

max_row = ws.max_row
count = 0

for row in range(2, max_row + 1):
    address = ws[f'D{row}'].value
    number = ws[f'L{row}'].value

    if isinstance(address, str) and address.startswith('Ain') and isinstance(number, (int, float)) and number < 40:
        count += 1

print(count)
